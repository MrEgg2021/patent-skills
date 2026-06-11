#!/usr/bin/env python3
"""Convert invention-point-extraction JSON output to XLSX."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_COLUMNS = ["专利公开号", "专利标题", "技术问题", "技术方案"]


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if lines and lines[0].strip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        return "\n".join(lines).strip()
    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.S | re.I)
    return fenced.group(1).strip() if fenced else text


def load_payload(path: Path) -> list[dict[str, Any]]:
    raw = strip_code_fence(read_text(path))
    payload = json.loads(raw)
    if isinstance(payload, dict):
        if isinstance(payload.get("results"), list):
            payload = payload["results"]
        elif isinstance(payload.get("data"), list):
            payload = payload["data"]
        else:
            payload = [payload]
    if not isinstance(payload, list):
        raise ValueError("JSON must be an object, an array, or an object with results/data array")
    rows = [item for item in payload if isinstance(item, dict)]
    if not rows:
        raise ValueError("No object rows found")
    return rows


def numbered(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        items = [str(item).strip() for item in value if str(item).strip()]
        if len(items) == 1:
            return items[0]
        return "\n".join(f"{idx}. {item}" for idx, item in enumerate(items, 1))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def split_invention_points(text: str) -> tuple[list[str], list[str]]:
    problems: list[str] = []
    solutions: list[str] = []
    if not text:
        return problems, solutions
    chunks = re.split(r"\n\s*(?=\d+[\.、])", text.strip())
    for chunk in chunks:
        problem = ""
        solution = ""
        m_problem = re.search(r"技术问题[:：]\s*(.*?)(?=技术方案[:：]|$)", chunk, flags=re.S)
        m_solution = re.search(r"技术方案[:：]\s*(.*)$", chunk, flags=re.S)
        if m_problem:
            problem = m_problem.group(1).strip()
        if m_solution:
            solution = m_solution.group(1).strip()
        if not problem and not solution:
            problem = chunk.strip()
        if problem:
            problems.append(re.sub(r"^\d+[\.、]\s*", "", problem))
        if solution:
            solutions.append(re.sub(r"^\d+[\.、]\s*", "", solution))
    return problems, solutions


def normalize_row(item: dict[str, Any], fallback_name: str) -> dict[str, str]:
    pub_no = (
        item.get("专利公开号")
        or item.get("公开号")
        or item.get("publication_number")
        or item.get("文件名称")
        or item.get("file_name")
        or fallback_name
    )
    title = item.get("专利标题") or item.get("标题") or item.get("title") or "（未识别）"
    problems = item.get("技术问题") or item.get("technical_problem")
    solutions = item.get("技术方案") or item.get("technical_solution")

    error = item.get("解析失败") or item.get("error") or item.get("parse_error")
    if error and not problems and not solutions:
        problems = [f"解析失败：{error}"]
        solutions = [f"解析失败：{error}"]

    if (not problems or not solutions) and item.get("发明点"):
        parsed_problems, parsed_solutions = split_invention_points(numbered(item.get("发明点")))
        problems = problems or parsed_problems
        solutions = solutions or parsed_solutions

    return {
        "专利公开号": numbered(pub_no),
        "专利标题": numbered(title),
        "技术问题": numbered(problems),
        "技术方案": numbered(solutions),
    }


def autosize_and_wrap(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    for header in worksheet[1]:
        header.font = Font(bold=True)
        header.fill = header_fill
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {"A": 18, "B": 40, "C": 60, "D": 60}
    for letter, width in widths.items():
        worksheet.column_dimensions[letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == "A":
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert invention point JSON to xlsx")
    parser.add_argument("input", help="JSON file from LLM extraction")
    parser.add_argument("output", help="XLSX output path")
    parser.add_argument("--sheet-name", default="发明点解析")
    args = parser.parse_args()

    input_path = Path(args.input)
    rows = [normalize_row(item, input_path.stem) for item in load_payload(input_path)]
    frame = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=args.sheet_name)
        autosize_and_wrap(writer, args.sheet_name)
    print(f"saved: {output}")
    print(f"rows: {len(frame)}")


if __name__ == "__main__":
    sys.dont_write_bytecode = True
    main()
