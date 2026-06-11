#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

sys.dont_write_bytecode = True

try:
    from docx import Document
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"missing_dependency: python-docx ({exc})")

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"missing_dependency: openpyxl ({exc})")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


PUBLICATIONS = [
    "CN113487713A",
    "CN120219484A",
    "CN119687902B",
    "CN113009759A",
    "CN204119349U",
    "US20130176412A1",
    "WO2015186131A1",
    "KR200460570Y1",
    "EP3160114B1",
    "JP2006251217A",
]

EXPECTED = {
    "CN113487713A": {
        "source_office": "CN",
        "pdf_source": "google",
        "warnings": [],
        "drawing_total": 3,
        "drawing_selected": 3,
        "drawing_hd_selected": 3,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 3,
        "drawing_caption_source": "description-of-drawings",
        "inline_downloaded": 13,
        "description_block_min": 5,
    },
    "CN120219484A": {
        "source_office": "CN",
        "pdf_source": "google",
        "warnings": [
            "google_drawings_not_exposed_on_page",
            "no_drawings_found_after_full_and_thumbnail_attempt",
        ],
        "drawing_total": 0,
        "drawing_selected": 0,
        "drawing_hd_selected": 0,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 1,
        "drawing_caption_source": "description-of-drawings",
        "inline_downloaded": 0,
        "description_block_min": 5,
    },
    "CN119687902B": {
        "source_office": "CN",
        "pdf_source": "none",
        "warnings": [
            "google_drawings_not_exposed_on_page",
            "no_drawings_found_after_full_and_thumbnail_attempt",
            "google_pdf_not_found",
        ],
        "drawing_total": 0,
        "drawing_selected": 0,
        "drawing_hd_selected": 0,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 5,
        "drawing_caption_source": "description-of-drawings",
        "inline_downloaded": 0,
        "description_block_min": 10,
    },
    "CN113009759A": {
        "source_office": "CN",
        "pdf_source": "google",
        "warnings": [],
        "drawing_total": 5,
        "drawing_selected": 5,
        "drawing_hd_selected": 5,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 5,
        "drawing_caption_source": "description-of-drawings",
        "inline_downloaded": 0,
        "description_block_min": 5,
    },
    "CN204119349U": {
        "source_office": "CN",
        "pdf_source": "google",
        "warnings": [
            "google_drawings_not_exposed_on_page",
            "no_drawings_found_after_full_and_thumbnail_attempt",
        ],
        "drawing_total": 0,
        "drawing_selected": 0,
        "drawing_hd_selected": 0,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 6,
        "drawing_caption_source": "description-of-drawings",
        "inline_downloaded": 0,
        "description_block_min": 10,
        "description_any_markers": [
            "技术领域",
            "背景技术",
            "实用新型内容",
            "具体实施方式",
        ],
    },
    "US20130176412A1": {
        "source_office": "US",
        "pdf_source": "google",
        "warnings": [],
        "drawing_total": 27,
        "drawing_selected": 27,
        "drawing_hd_selected": 27,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 27,
        "drawing_caption_source": "description-of-drawings",
        "description_block_min": 10,
        "required_heading_phrases": [
            "FIELD OF THE INVENTION",
            "BACKGROUND OF THE INVENTION",
            "SUMMARY OF THE INVENTION",
        ],
    },
    "WO2015186131A1": {
        "source_office": "WO",
        "pdf_source": "google",
        "warnings": [
            "google_drawings_not_exposed_on_page",
            "no_drawings_found_after_full_and_thumbnail_attempt",
        ],
        "drawing_total": 0,
        "drawing_selected": 0,
        "drawing_hd_selected": 0,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 7,
        "drawing_caption_source": "heading-following-list",
        "description_block_min": 10,
    },
    "KR200460570Y1": {
        "source_office": "KR",
        "pdf_source": "google",
        "warnings": [],
        "drawing_total": 7,
        "drawing_selected": 7,
        "drawing_hd_selected": 7,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 7,
        "drawing_caption_source": "description-of-drawings",
        "description_block_min": 10,
    },
    "EP3160114B1": {
        "source_office": "EP",
        "pdf_source": "google",
        "warnings": [
            "abstract_not_exposed_on_page",
            "drawing_caption_image_count_mismatch:7:4",
        ],
        "drawing_total": 4,
        "drawing_selected": 4,
        "drawing_hd_selected": 4,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 7,
        "drawing_caption_source": "heading-following-list",
        "description_block_min": 10,
        "required_heading_phrases": [
            "TECHNICAL FIELD",
            "BACKGROUND",
            "SUMMARY",
            "BRIEF DESCRIPTION OF DRAWINGS",
        ],
        "abstract_must_be_empty": True,
    },
    "JP2006251217A": {
        "source_office": "JP",
        "pdf_source": "google",
        "warnings": [],
        "drawing_total": 8,
        "drawing_selected": 8,
        "drawing_hd_selected": 8,
        "drawing_thumbnail_fallback": 0,
        "drawing_caption_total": 8,
        "drawing_caption_source": "description-of-drawings",
        "description_block_min": 10,
    },
}

FORBIDDEN_TOKENS = [
    "[IMAGE_MISSING]",
    "[INLINE_IMAGE_MISSING]",
    "URL: ",
    "\u2011",
    "Global patent litigation dataset",
    "Darts-ip",
    "Creative Commons Attribution",
]

DOCX_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}
TEXT_TABLE_FILENAME = "patent_text_records.xlsx"
TEXT_TABLE_SHEET = "patent_texts"
TEXT_TABLE_LIMIT = 32000


def run_export(script_path: Path, output_dir: Path, proxy_url: str, timeout_sec: int) -> None:
    cmd = [
        sys.executable,
        str(script_path),
        "--mode",
        "patent",
        "--publication-numbers",
        ",".join(PUBLICATIONS),
        "--timeout-sec",
        str(timeout_sec),
        "--output-dir",
        str(output_dir),
    ]
    if proxy_url.strip():
        cmd.extend(["--proxy-url", proxy_url])
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
    if proc.returncode != 0:
        raise AssertionError(
            "export_artifacts_failed\n"
            f"command={' '.join(cmd)}\n"
            f"stdout={proc.stdout}\n"
            f"stderr={proc.stderr}"
        )


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_docx_text(path: Path) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def parse_docx_paragraphs(path: Path) -> list[dict[str, object]]:
    with ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    rows: list[dict[str, object]] = []
    for p in root.findall(".//w:p", DOCX_NS):
        text = "".join(t.text or "" for t in p.findall(".//w:t", DOCX_NS))
        images = len(p.findall(".//a:blip", DOCX_NS))
        rows.append({"text": text.strip(), "images": images})
    return rows


def truncate_text_for_table(text: str) -> str:
    return text[:TEXT_TABLE_LIMIT] if len(text) > TEXT_TABLE_LIMIT else text


def load_text_table_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        if TEXT_TABLE_SHEET not in workbook.sheetnames:
            raise AssertionError(f"text table missing sheet: {TEXT_TABLE_SHEET}")
        worksheet = workbook[TEXT_TABLE_SHEET]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    headers = ["" if cell is None else str(cell) for cell in rows[0]]
    items: list[dict[str, object]] = []
    for raw_row in rows[1:]:
        items.append({headers[idx]: raw_row[idx] for idx in range(len(headers))})
    return items


def next_nonempty(rows: list[dict[str, object]], start: int) -> int:
    for idx in range(start, len(rows)):
        if rows[idx]["text"] or rows[idx]["images"]:
            return idx
    return -1


def assert_no_forbidden_tokens(text: str, where: str, failures: list[str]) -> None:
    for token in FORBIDDEN_TOKENS:
        if token in text:
            failures.append(f"{where} contains forbidden token: {token}")


def require_equal(actual: object, expected: object, label: str, failures: list[str]) -> None:
    if actual != expected:
        failures.append(f"{label} expected {expected!r} got {actual!r}")


def assert_text_table(base_dir: Path, failures: list[str]) -> None:
    table_path = base_dir / TEXT_TABLE_FILENAME
    if not table_path.exists():
        failures.append(f"missing text table: {table_path}")
        return

    try:
        rows = load_text_table_rows(table_path)
    except Exception as exc:
        failures.append(f"text table unreadable: {exc}")
        return

    if len(rows) != len(PUBLICATIONS):
        failures.append(f"text table row count expected {len(PUBLICATIONS)} got {len(rows)}")

    row_by_publication: dict[str, dict[str, object]] = {}
    for row in rows:
        publication_number = str(row.get("publication_number") or "").strip()
        if not publication_number:
            failures.append("text table contains blank publication_number")
            continue
        if publication_number in row_by_publication:
            failures.append(f"text table contains duplicate publication_number: {publication_number}")
            continue
        row_by_publication[publication_number] = row

    missing = [publication for publication in PUBLICATIONS if publication not in row_by_publication]
    if missing:
        failures.append(f"text table missing publications: {missing!r}")

    for publication_number in ("US20130176412A1", "CN113487713A"):
        row = row_by_publication.get(publication_number)
        if row is None:
            continue
        parsed = load_json(base_dir / publication_number / "parsed.json")
        sections = parsed.get("sections", {})
        for section_key in ("abstract", "claims", "description"):
            expected_text = truncate_text_for_table(str(sections.get(section_key, "")))
            actual_text = str(row.get(section_key) or "")
            if actual_text != expected_text:
                failures.append(
                    f"text table {publication_number} {section_key} mismatch with parsed.json"
                )
        if str(row.get("crawl_status") or "") != "success":
            failures.append(f"text table {publication_number} crawl_status expected 'success'")
        docx_path = str(row.get("docx_path") or "")
        pdf_path = str(row.get("pdf_path") or "")
        if not docx_path.endswith(f"{publication_number}.docx"):
            failures.append(f"text table {publication_number} docx_path mismatch: {docx_path!r}")
        if not pdf_path.endswith(f"{publication_number}.pdf"):
            failures.append(f"text table {publication_number} pdf_path mismatch: {pdf_path!r}")


def assert_metadata(publication_number: str, base_dir: Path, failures: list[str]) -> None:
    expected = EXPECTED[publication_number]
    patent_dir = base_dir / publication_number
    run_metadata = load_json(patent_dir / "run_metadata.json")
    parsed = load_json(patent_dir / "parsed.json")
    image_manifest = load_json(patent_dir / "image_manifest.json")
    docx_path = patent_dir / f"{publication_number}.docx"
    parsed_text = json.dumps(parsed, ensure_ascii=False)
    docx_text = load_docx_text(docx_path)
    docx_rows = parse_docx_paragraphs(docx_path)
    description_text = str(parsed.get("sections", {}).get("description", ""))
    abstract_text = str(parsed.get("sections", {}).get("abstract", ""))
    section_stats = parsed.get("section_stats", {})
    description_stats = section_stats.get("description", {})

    assert_no_forbidden_tokens(parsed_text, f"{publication_number}/parsed.json", failures)
    assert_no_forbidden_tokens(docx_text, f"{publication_number}.docx", failures)

    require_equal(parsed.get("source_office"), expected["source_office"], f"{publication_number} parsed source_office", failures)
    require_equal(run_metadata.get("source_office"), expected["source_office"], f"{publication_number} metadata source_office", failures)
    require_equal(run_metadata.get("pdf_source"), expected["pdf_source"], f"{publication_number} metadata pdf_source", failures)
    require_equal(run_metadata.get("drawing_total"), expected["drawing_total"], f"{publication_number} metadata drawing_total", failures)
    require_equal(run_metadata.get("drawing_selected"), expected["drawing_selected"], f"{publication_number} metadata drawing_selected", failures)
    require_equal(run_metadata.get("drawing_hd_selected"), expected["drawing_hd_selected"], f"{publication_number} metadata drawing_hd_selected", failures)
    require_equal(run_metadata.get("drawing_thumbnail_fallback"), expected["drawing_thumbnail_fallback"], f"{publication_number} metadata drawing_thumbnail_fallback", failures)
    require_equal(parsed.get("drawing_caption_total"), expected["drawing_caption_total"], f"{publication_number} parsed drawing_caption_total", failures)
    require_equal(run_metadata.get("drawing_caption_total"), expected["drawing_caption_total"], f"{publication_number} metadata drawing_caption_total", failures)
    require_equal(parsed.get("drawing_caption_source"), expected["drawing_caption_source"], f"{publication_number} parsed drawing_caption_source", failures)
    require_equal(run_metadata.get("drawing_caption_source"), expected["drawing_caption_source"], f"{publication_number} metadata drawing_caption_source", failures)

    if sorted(run_metadata.get("warnings", [])) != sorted(expected["warnings"]):
        failures.append(
            f"{publication_number} warnings expected {expected['warnings']!r} got {run_metadata.get('warnings', [])!r}"
        )

    block_count = int(description_stats.get("block_count", 0) or 0)
    if block_count < int(expected.get("description_block_min", 1)):
        failures.append(
            f"{publication_number} description block_count expected >= {expected['description_block_min']} got {block_count}"
        )

    if expected.get("abstract_must_be_empty") and abstract_text.strip():
        failures.append(f"{publication_number} abstract expected to be empty but was populated")

    for marker in expected.get("required_heading_phrases", []):
        if marker not in description_text and marker not in docx_text:
            failures.append(f"{publication_number} missing required heading phrase: {marker}")

    any_markers = expected.get("description_any_markers", [])
    if any_markers and not any(marker in description_text for marker in any_markers):
        failures.append(f"{publication_number} description missing any of expected markers: {any_markers!r}")

    selected = [item for item in image_manifest if item.get("selected_for_drawings")]
    if len(selected) != expected["drawing_selected"]:
        failures.append(
            f"{publication_number} selected drawing count expected {expected['drawing_selected']} got {len(selected)}"
        )

    drawing_hd_selected = sum(1 for item in selected if item.get("image_role") == "drawing_hd")
    if drawing_hd_selected != expected["drawing_hd_selected"]:
        failures.append(
            f"{publication_number} selected HD drawing count expected {expected['drawing_hd_selected']} got {drawing_hd_selected}"
        )

    if "inline_downloaded" in expected:
        inline_downloaded = sum(
            1
            for item in image_manifest
            if item.get("image_role") == "inline_formula" and item.get("download_status") == "downloaded"
        )
        if inline_downloaded != expected["inline_downloaded"]:
            failures.append(
                f"{publication_number} downloaded inline image count expected {expected['inline_downloaded']} got {inline_downloaded}"
            )
    else:
        inline_downloaded = sum(
            1
            for item in image_manifest
            if item.get("image_role") == "inline_formula" and item.get("download_status") == "downloaded"
        )

    drawing_heading_index = -1
    warnings_heading_index = len(docx_rows)
    for idx, row in enumerate(docx_rows):
        if row["text"] == "附图说明":
            drawing_heading_index = idx
        if row["text"] == "Warnings":
            warnings_heading_index = idx
            break

    if drawing_heading_index < 0:
        failures.append(f"{publication_number}.docx missing 附图说明 heading")
        return

    body_image_count = sum(
        1 for row in docx_rows[:drawing_heading_index] if row["images"] and not row["text"]
    )
    if body_image_count != inline_downloaded:
        failures.append(
            f"{publication_number}.docx body image count expected {inline_downloaded} got {body_image_count}"
        )

    if expected["drawing_selected"] == 0:
        drawing_images = sum(1 for row in docx_rows[drawing_heading_index + 1 : warnings_heading_index] if row["images"])
        if drawing_images != 0:
            failures.append(f"{publication_number}.docx expected 0 drawing images got {drawing_images}")
        return

    cursor = drawing_heading_index + 1
    for item in sorted(selected, key=lambda row: row.get("sequence_no", 0)):
        seq = item.get("sequence_no", 0)
        image_idx = next_nonempty(docx_rows, cursor)
        if image_idx < 0:
            failures.append(f"{publication_number}.docx missing drawing image paragraph for 图{seq}")
            break
        if not docx_rows[image_idx]["images"]:
            failures.append(
                f"{publication_number}.docx drawing group 图{seq} does not start with an image paragraph"
            )
            break

        caption_idx = next_nonempty(docx_rows, image_idx + 1)
        if caption_idx < 0:
            failures.append(f"{publication_number}.docx missing drawing caption paragraph for 图{seq}")
            break
        caption_text = str(docx_rows[caption_idx]["text"])
        if not caption_text.startswith(f"图{seq}"):
            failures.append(
                f"{publication_number}.docx drawing caption order mismatch for 图{seq}: got {caption_text!r}"
            )
            break
        cursor = caption_idx + 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Google Patent regression crawl and validate outputs.")
    parser.add_argument("--output-dir", required=True, help="Output directory for crawl artifacts.")
    parser.add_argument("--proxy-url", default="", help="Optional proxy URL for online crawl.")
    parser.add_argument("--timeout-sec", type=int, default=30, help="HTTP timeout for patent fetches.")
    parser.add_argument(
        "--skip-export",
        action="store_true",
        help="Validate an existing output directory without invoking export_artifacts.py.",
    )
    args = parser.parse_args()

    script_path = Path(__file__).with_name("export_artifacts.py")
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not args.skip_export:
        run_export(script_path, output_dir, args.proxy_url, args.timeout_sec)

    failures: list[str] = []
    for publication_number in PUBLICATIONS:
        patent_dir = output_dir / publication_number
        if not patent_dir.exists():
            failures.append(f"missing publication output directory: {patent_dir}")
            continue
        assert_metadata(publication_number, output_dir, failures)
    assert_text_table(output_dir, failures)

    result = {
        "ok": not failures,
        "output_dir": str(output_dir),
        "publication_numbers": PUBLICATIONS,
        "failures": failures,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    raise SystemExit(0 if not failures else 1)


if __name__ == "__main__":
    main()
