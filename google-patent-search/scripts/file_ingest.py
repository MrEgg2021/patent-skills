#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

sys.dont_write_bytecode = True

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

SKILL_NAME = 'google-patent-search'
RESERVED_KEYS = {
    'skill_name',
    'timestamp',
    'required_inputs',
    'stage_plan',
    'validation_errors',
    'ready',
    'parsed_artifacts',
    'context_for_prompt_runtime',
    'warnings',
    'search_payload',
}
PUBLICATION_RE = re.compile(r'^[A-Za-z]{2}(?=.*\d)[A-Za-z0-9./-]+$')
PATENT_URL_RE = re.compile(r'https?://patents\.google\.com/patent/([A-Za-z0-9./-]+)', re.IGNORECASE)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_text_flexible(path: Path) -> str:
    for encoding in ('utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'gb18030'):
        try:
            return path.read_text(encoding=encoding)
        except Exception:
            continue
    raise UnicodeDecodeError('codec', b'', 0, 1, f'cannot decode {path}')


def load_json(path: Path) -> Any:
    return json.loads(read_text_flexible(path))


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ''
    if isinstance(value, (list, tuple, set)):
        return len(value) == 0
    return False


def as_str(value: Any) -> str:
    return '' if value is None else str(value).strip()


def as_int(value: Any, default: int | None = None) -> int | None:
    if value is None:
        return default
    if isinstance(value, int):
        return value
    text = as_str(value)
    if not text:
        return default
    try:
        return int(text)
    except Exception:
        return default


def as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    token = as_str(value).lower()
    if token in {'1', 'true', 'yes', 'y', 'on'}:
        return True
    if token in {'0', 'false', 'no', 'n', 'off'}:
        return False
    return default


def as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = as_str(value)
    if not text:
        return []
    if text.startswith('[') and text.endswith(']'):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if str(v).strip()]
        except Exception:
            pass
    return [part.strip() for part in re.split(r'[\n,;]+', text) if part.strip()]


def resolve_path(raw: Any) -> Path:
    path = Path(str(raw)).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    return path


def normalize_answers(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    out: dict[str, Any] = {}
    required_inputs = payload.get('required_inputs')
    if isinstance(required_inputs, dict):
        out.update(required_inputs)

    stage_plan = payload.get('stage_plan')
    if isinstance(stage_plan, list):
        for stage in stage_plan:
            if not isinstance(stage, dict):
                continue
            key = stage.get('option_key')
            value = stage.get('selected_option')
            if isinstance(key, str) and key and value is not None:
                out.setdefault(key, value)

    for key, value in payload.items():
        if key in RESERVED_KEYS:
            continue
        out[key] = value
    return out


def load_inputs(args: argparse.Namespace) -> tuple[dict[str, Any], dict[str, Any]]:
    answers: dict[str, Any] = {}
    context: dict[str, Any] = {}

    if args.answers:
        answers.update(normalize_answers(json.loads(args.answers)))
    if args.answers_file:
        answers.update(normalize_answers(load_json(Path(args.answers_file))))

    if args.context:
        raw_context = json.loads(args.context)
        if isinstance(raw_context, dict):
            context.update(raw_context)
    if args.context_file:
        file_context = load_json(Path(args.context_file))
        if isinstance(file_context, dict):
            context.update(file_context)

    return answers, context


def headers_rows(raw_rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    first_non_blank = None
    for idx, row in enumerate(raw_rows):
        if any(as_str(cell) for cell in row):
            first_non_blank = idx
            break
    if first_non_blank is None:
        return [], []

    headers: list[str] = []
    for idx, cell in enumerate(raw_rows[first_non_blank]):
        label = as_str(cell)
        headers.append(label if label else f'column_{idx + 1}')

    rows: list[list[str]] = []
    width = len(headers)
    for row in raw_rows[first_non_blank + 1 :]:
        values = [as_str(cell) for cell in row[:width]]
        if len(values) < width:
            values.extend([''] * (width - len(values)))
        if any(values):
            rows.append(values)
    return headers, rows


def read_csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    for encoding in ('utf-8-sig', 'gb18030'):
        try:
            with path.open('r', encoding=encoding, newline='') as handle:
                rows = [[as_str(cell) for cell in row] for row in csv.reader(handle)]
            return headers_rows(rows)
        except Exception:
            continue
    raise ValueError('csv_read_failed')


def read_xlsx_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError('missing_dependency:openpyxl') from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [[as_str(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return headers_rows(rows)


def pick_publication_col(headers: list[str]) -> int:
    if not headers:
        return 0
    tokens = (
        'publication',
        'publication_number',
        'patent',
        'pn',
        '公开',
        '公开号',
        '文献',
        '申请号',
    )
    for idx, header in enumerate(headers):
        normalized = as_str(header).lower().replace(' ', '_')
        if any(token in normalized for token in tokens):
            return idx
    return 0


def normalize_publications(raw_values: list[str]) -> tuple[list[str], list[str]]:
    publications: list[str] = []
    warnings: list[str] = []
    seen: set[str] = set()

    for value in raw_values:
        text = as_str(value)
        if not text or text in seen:
            continue
        seen.add(text)
        if PUBLICATION_RE.fullmatch(text) is None:
            warnings.append(f'publication_number:format_unverified:{text}')
        publications.append(text)
    return publications, warnings


def extract_publication_candidates(node: Any) -> list[str]:
    candidates: list[str] = []
    if isinstance(node, dict):
        for key in ('publication_number', 'publicationNumber', 'publication', 'patent_number', 'patentNumber'):
            value = node.get(key)
            if value not in (None, '', []):
                candidates.extend(as_list(value))
        url_value = node.get('url') or node.get('link') or node.get('href')
        if isinstance(url_value, str):
            match = PATENT_URL_RE.search(url_value)
            if match:
                candidates.append(match.group(1))
        for value in node.values():
            candidates.extend(extract_publication_candidates(value))
    elif isinstance(node, list):
        for item in node:
            candidates.extend(extract_publication_candidates(item))
    elif isinstance(node, str):
        match = PATENT_URL_RE.search(node)
        if match:
            candidates.append(match.group(1))
        elif PUBLICATION_RE.fullmatch(node.strip()):
            candidates.append(node.strip())
    return candidates


def parse_single_mode(answers: dict[str, Any], errors: list[str], warnings: list[str]) -> tuple[list[str], dict[str, Any]]:
    values = as_list(answers.get('publication_number'))
    if not values:
        errors.append('publication_number:required')
        return [], {}

    publications, local_warnings = normalize_publications(values)
    warnings.extend(local_warnings)
    if not publications:
        errors.append('publication_number:no_valid_entries')
        return [], {}

    artifact = {
        'source': 'inline',
        'input_count': len(values),
        'publication_count': len(publications),
    }
    return publications, artifact


def parse_batch_mode(answers: dict[str, Any], errors: list[str], warnings: list[str]) -> tuple[list[str], dict[str, Any]]:
    raw = answers.get('patent_list_file')
    if is_blank(raw):
        errors.append('patent_list_file:required')
        return [], {}

    path = resolve_path(raw)
    if not path.exists() or not path.is_file():
        errors.append(f'patent_list_file:file_not_found:{path}')
        return [], {}

    extension = path.suffix.lower()
    if extension == '.xls':
        errors.append('patent_list_file:unsupported_legacy_xls_convert_to_xlsx')
        return [], {}
    if extension not in {'.csv', '.xlsx'}:
        errors.append(f'patent_list_file:unsupported_file_extension:{extension or "none"}')
        return [], {}

    try:
        if extension == '.csv':
            headers, rows = read_csv_rows(path)
            parser_name = 'csv'
        else:
            headers, rows = read_xlsx_rows(path)
            parser_name = 'openpyxl'
    except RuntimeError as exc:
        errors.append(f'patent_list_file:{exc}')
        return [], {}
    except ValueError as exc:
        errors.append(f'patent_list_file:{exc}')
        return [], {}
    except Exception as exc:
        errors.append(f'patent_list_file:parse_failed:{exc}')
        return [], {}

    column_index = pick_publication_col(headers)
    raw_values = [row[column_index] for row in rows if column_index < len(row)]
    publications, local_warnings = normalize_publications(raw_values)
    warnings.extend(local_warnings)
    if not publications:
        errors.append('patent_list_file:no_publication_numbers_extracted')
        return [], {}

    artifact = {
        'path': str(path),
        'parser_name': parser_name,
        'row_count': len(rows),
        'column_count': len(headers),
        'publication_column_index': column_index,
        'publication_column_name': headers[column_index] if headers and column_index < len(headers) else '',
        'publication_count': len(publications),
    }
    return publications, artifact


def parse_results_json_mode(answers: dict[str, Any], errors: list[str], warnings: list[str]) -> tuple[list[str], dict[str, Any]]:
    raw = answers.get('results_json_file')
    if is_blank(raw):
        errors.append('results_json_file:required')
        return [], {}

    path = resolve_path(raw)
    if not path.exists() or not path.is_file():
        errors.append(f'results_json_file:file_not_found:{path}')
        return [], {}
    if path.suffix.lower() != '.json':
        errors.append(f'results_json_file:unsupported_file_extension:{path.suffix.lower() or "none"}')
        return [], {}

    try:
        payload = load_json(path)
    except Exception as exc:
        errors.append(f'results_json_file:parse_failed:{exc}')
        return [], {}

    raw_candidates = extract_publication_candidates(payload)
    publications, local_warnings = normalize_publications(raw_candidates)
    warnings.extend(local_warnings)
    if not publications:
        errors.append('results_json_file:no_publication_numbers_extracted')
        return [], {}

    source_results = payload.get('results') if isinstance(payload, dict) else payload
    artifact = {
        'path': str(path),
        'source_result_count': len(source_results) if isinstance(source_results, list) else 0,
        'publication_count': len(publications),
        'publication_numbers_preview': publications[:10],
    }
    return publications, artifact


def resolve_mode(answers: dict[str, Any]) -> str:
    mode = as_str(answers.get('mode')).lower()
    if mode:
        return mode
    return 'detail' if any(not is_blank(answers.get(key)) for key in ('detail_input_mode', 'publication_number', 'patent_list_file', 'results_json_file')) else ''


def resolve_detail_input_mode(answers: dict[str, Any]) -> str:
    mode = as_str(answers.get('detail_input_mode')).lower()
    if mode in {'single', 'batch', 'results_json'}:
        return mode
    has_single = not is_blank(answers.get('publication_number'))
    has_batch = not is_blank(answers.get('patent_list_file'))
    has_results = not is_blank(answers.get('results_json_file'))
    if has_single and not has_batch and not has_results:
        return 'single'
    if has_batch and not has_single and not has_results:
        return 'batch'
    if has_results and not has_single and not has_batch:
        return 'results_json'
    return ''


def build_required_inputs(answers: dict[str, Any], detail_input_mode: str) -> dict[str, Any]:
    required_inputs: dict[str, Any] = {
        'mode': 'detail',
        'project_id': answers.get('project_id'),
        'detail_input_mode': detail_input_mode,
        'file_formats': answers.get('file_formats') or 'both',
        'proxy_url': answers.get('proxy_url') or '',
        'request_timeout_sec': as_int(answers.get('request_timeout_sec'), 30),
        'use_hd_drawings': as_bool(answers.get('use_hd_drawings'), True),
        'include_abstract': as_bool(answers.get('include_abstract'), True),
        'include_claims': as_bool(answers.get('include_claims'), True),
        'include_description': as_bool(answers.get('include_description'), True),
        'include_drawings': as_bool(answers.get('include_drawings'), True),
    }
    if detail_input_mode == 'single':
        required_inputs['publication_number'] = answers.get('publication_number')
    elif detail_input_mode == 'batch':
        required_inputs['patent_list_file'] = answers.get('patent_list_file')
    elif detail_input_mode == 'results_json':
        required_inputs['results_json_file'] = answers.get('results_json_file')
    return required_inputs


def main() -> None:
    parser = argparse.ArgumentParser(description='Skill-local file ingestion for google-patent-search detail mode')
    parser.add_argument('--answers')
    parser.add_argument('--answers-file')
    parser.add_argument('--context')
    parser.add_argument('--context-file')
    parser.add_argument('--context-only', action='store_true')
    parser.add_argument('--non-interactive', action='store_true')
    args = parser.parse_args()

    answers, context = load_inputs(args)
    validation_errors: list[str] = []
    warnings: list[str] = []
    parsed_artifacts: dict[str, Any] = {}

    mode = resolve_mode(answers)
    if mode != 'detail':
        validation_errors.append('mode:detail_required')

    detail_input_mode = resolve_detail_input_mode(answers)
    if not detail_input_mode:
        validation_errors.append('detail_input_mode:required')

    publications: list[str] = []
    if detail_input_mode == 'single':
        publications, artifact = parse_single_mode(answers, validation_errors, warnings)
        if artifact:
            parsed_artifacts['publication_number'] = artifact
    elif detail_input_mode == 'batch':
        publications, artifact = parse_batch_mode(answers, validation_errors, warnings)
        if artifact:
            parsed_artifacts['patent_list_file'] = artifact
    elif detail_input_mode == 'results_json':
        publications, artifact = parse_results_json_mode(answers, validation_errors, warnings)
        if artifact:
            parsed_artifacts['results_json_file'] = artifact

    if publications:
        parsed_artifacts['publication_numbers'] = {
            'count': len(publications),
            'items': publications,
        }

    runtime_context = dict(context)
    runtime_context.update(
        {
            'mode': 'detail',
            'detail_input_mode': detail_input_mode,
            'publication_numbers': publications,
            'publication_numbers_text': '\n'.join(publications),
            'total_count': len(publications),
            'file_formats': answers.get('file_formats') or 'both',
            'proxy_url': answers.get('proxy_url') or '',
            'request_timeout_sec': as_int(answers.get('request_timeout_sec'), 30),
            'use_hd_drawings': as_bool(answers.get('use_hd_drawings'), True),
            'include_abstract': as_bool(answers.get('include_abstract'), True),
            'include_claims': as_bool(answers.get('include_claims'), True),
            'include_description': as_bool(answers.get('include_description'), True),
            'include_drawings': as_bool(answers.get('include_drawings'), True),
            'patent_base_url_template': 'https://patents.google.com/patent/{publication_number}',
            'require_online_access': True,
            'en_suffix_disabled': True,
        }
    )

    result = {
        'skill_name': SKILL_NAME,
        'timestamp': now_iso(),
        'required_inputs': build_required_inputs(answers, detail_input_mode),
        'parsed_artifacts': parsed_artifacts,
        'context_for_prompt_runtime': runtime_context,
        'validation_errors': sorted(set(validation_errors)),
        'warnings': sorted(set(warnings)),
        'ready': len(set(validation_errors)) == 0,
    }

    if args.context_only:
        print(json.dumps(runtime_context, ensure_ascii=False, indent=2))
        return

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
